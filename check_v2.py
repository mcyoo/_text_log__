import openpyxl
from openpyxl.styles import PatternFill, Color, Font
import codecs
import os

# 매개변수 file 안에 hostname을 반환하는 함수
def name(file):
	name = 0
	file.seek(0)
	while True:
		line=file.readline()
		hop = line.count('#')
		if not line:
			break
		if hop == 0 or hop > 5: continue 
		else :
			temp = line
			for i in range(0,hop):
				name+=temp.find('#')
				name+=1
				temp=line[name:]
				i+=1

			line=line[:name-1]
			break

	return line

#엑셀 문서를 만드는 함수
def excelOpen():
	try:
		wb = openpyxl.Workbook()
		return wb

	except Exception as ex: # 에러 종류
		print('error', ex)
		return -1

# 매개변수 file 안에 match 가 되는 내용을 리스트 형태로 반환하는 함수
def data_crawling(file,name):
	log = []
	log.clear()
	file.seek(0)
	while True:
		line=file.readline()
		#tem1=line.find('Temperature normal: ')
		tem2=line.find('five minutes: ')
		tem3=line.find('Processor Pool Total:')
		tem4=line.find('show redun')
		tem5=line.find('Free: ')
		#tem5=line.find('air inlet')
		#tem6=line.find('air outlet')
		tem7=line.find('20G')
		tem8=line.find('crashinfo_201911')
		tem9=line.find('crashinfo_201912')
		#tem10=line.find('Image Version =')
		#tem11=line.find('inlet temperature:')
		#tem12=line.find('outlet temperature:')
		#tem13=line.find('Temp:   outlet')
		#tem14=line.find('Temp:    inlet')
		#tem15=line.find('STANDBY HOT')
		tem16=line.find('system-report_RP')#9407 crashinfo
		tem17=line.find('pxf_crashinfo_201911')
		tem18=line.find('pxf_crashinfo_201912')
		tem19=line.find('OFFLINE')
		#tem20=line.find('Power Entry Module 0 type AC status:')
		#tem21=line.find('Power Entry Module 1 type AC status:')
		#tem22=line.find('Power supplies currently available')
		tem23=line.find('CPU% per minute (last 60 minutes)')
		tem24=line.find('show module')
		tem25=line.find('show power')
		tem26=line.find('show ip ospf nei')
		tem27=line.find('show controllers clock-reference')
		tem28=line.find('show env')
		tem29=line.find('show hw bay all oir')
		
		#if tem1 >= 0:
		#	log.append(line)
		if tem2 >= 0:
			log.append(line)
		if tem3 >= 0:
			log.append(line)
		if tem4 >= 0:
			tell=file.tell()
			while True:
				line=file.readline()
				if line.find(name+'#') >= 0:
					break
				else:
					log.append(line)

				if not line:
					break
			file.seek(tell)

		if tem5 >= 0:
			log.append(line)
		#if tem6 >= 0:
		#	log.append(line)
		if tem7 >= 0:
			log.append(line)
		if tem8 >= 0:
			log.append(line)
		if tem9 >= 0:
			log.append(line)
		#if tem10 >= 0:
		#	log.append(line)
		#if tem11 >= 0:
		#	log.append(line)
		#if tem12 >= 0:
		#	log.append(line)
		#if tem13 >= 0:
		#	log.append(line)
		#if tem14 >= 0:
		#	log.append(line)
		#if tem15 >= 0:
		#	log.append(line)
		if tem16 >= 0:
			log.append(line)
		if tem17 >= 0:
			log.append(line)
		if tem18 >= 0:
			log.append(line)
		if tem19 >= 0:
			log.append(line)
		#if tem20 >= 0:
		#	log.append(line)
		#if tem21 >= 0:
		#	log.append(line)
		#if tem22 >= 0:
		#	log.append(line)
		if tem23 >= 0:
			tell=file.tell()
			while True:
				line=file.readline()
				if line.find(name+'#') >= 0:
					break
				else:
					log.append(line)

				if not line:
					break

			file.seek(tell)

		if tem24 >= 0:
			tell=file.tell()
			while True:
				line=file.readline()
				if line.find(name+'#') >= 0:
					break
				else:
					log.append(line)

				if not line:
					break
			file.seek(tell)

		if tem25 >= 0:
			tell=file.tell()
			while True:
				line=file.readline()
				if line.find(name+'#') >= 0:
					break
				else:
					log.append(line)

				if not line:
					break
			file.seek(tell)

		if tem26 >= 0:
			tell=file.tell()
			log.append('업링크 갯수(show ip ospf neighbor)')
			while True:
				line=file.readline()
				if line.find(name+'#') >= 0:
					break
				else:
					log.append(line)

				if not line:
					break
			file.seek(tell)

		if tem27 >= 0:
			tell=file.tell()
			while True:
				line=file.readline()
				if line.find(name+'#') >= 0:
					break
				else:
					log.append(line)

				if not line:
					break
			file.seek(tell)

		if tem28 >= 0:
			tell=file.tell()
			while True:
				line=file.readline()
				if line.find(name+'#') >= 0:
					break
				else:
					log.append(line)

				if not line:
					break
			file.seek(tell)

		if tem29 >= 0:
			tell=file.tell()
			while True:
				line=file.readline()
				if line.find(name+'#') >= 0:
					break
				else:
					log.append(line)

				if not line:
					break
			file.seek(tell)


		if not line:
			break

	return log

# 매개변수 file 안에 show log 를 찾아서, show log 안에 Aug, Sep 
def log_crawling(file,name):
	log = []
	log.clear()
	file.seek(0)
	while True:
		line=file.readline()
		tem=line.find('show log')
		if tem >= 0:
			while True:
				line1=file.readline()
				tem1=line1.find('Nov') # <-- 11월
				tem2=line1.find('Dec') # <-- 12월
				tem3=line1.find(name+'#')
				tem4=line1.find('UBR10000-5-UNREGSIDTIMEOUT:') #<- 이 로그 제외 
				tem5=line1.find('UBR10000-4-BADTXOFFSET:') #<- 이 로그 제외 
				tem6=line1.find('LICENSE-6-VIOLATION:') #<- 이 로그 제외 

				if tem1 >= 0 or tem2 >= 0:
					if tem4 >= 0 or tem5 >= 0 or tem6 >=0:
						pass
					else:
						log.append(line1)

				if tem3 >= 0:
					break

				if not line1:
					break

			break

		if not line:
			break

	return log

#pathnote 경로 안에 .txt .log 로 끝나는 파일을 열어서 리스트에 문자열을 하나씩 엑셀에 저장하는 함수
def LOG_crawling(wb,pathnote):
	
	x = 1
	rns_number = 0
	cmts_number = 0

	ws = wb.create_sheet('LOG')

	for (path, dir, files) in os.walk(pathnote):
	
		for filename in files:
			ext = os.path.splitext(filename)[-1]
			if ext == '.txt' or ext == '.log':
				print("%s/%s" % (path, filename))

				f = codecs.open("%s/%s" % (path, filename), 'r', "utf-8-sig", errors='ignore')

				#메모장에서 이름 가져오기
				namee = name(f)

				#10K 는 로그 저장 X 
				if namee.find('10K') >= 0 or namee.find('10k') >= 0 or namee.find('cBR8') >= 0 or namee.find('CMTS') >= 0 or namee.find('Anyang') >= 0 or namee.find('RFGW') >= 0:
					if namee.find('6509') >= 0 or namee.find('4507') >= 0: #안양 hostname 예외 경우 
						rns_number += 1
					
					else:
						cmts_number += 1
						#continue

				#R&S 개수 세기 
				else:
					rns_number += 1

				log = log_crawling(f,namee)

				ws.cell(x,1).value = namee
				ws.cell(x,1).font = Font(size=15, bold=True)

				for i in log:
					x+=1
					ws.cell(x,1).value = i

				x+=1

				f.close()

	#부가 기능
	print("R&S 개수 : %d"%rns_number)
	print("10K 개수 : %d"%cmts_number)


#pathnote 경로 안에 .txt .log 로 끝나는 파일을 열어서 리스트에 문자열을 하나씩 엑셀에 저장하는 함수
def check_crawring(wb,pathnote):
	
	x = 1
	rns_number = 0
	cmts_number = 0

	ws = wb.active # active 시트
	ws.title = '수집 데이터'

	for (path, dir, files) in os.walk(pathnote):
	
		for filename in files:
			ext = os.path.splitext(filename)[-1]
			if ext == '.txt' or ext == '.log':
				print("%s/%s" % (path, filename))

				f = codecs.open("%s/%s" % (path, filename), 'r', "utf-8-sig", errors='ignore')

				#메모장에서 이름 가져오기
				namee = name(f)
				
				# CMTS 개수 세기
				if namee.find('10K') >= 0 or namee.find('10k') >= 0 or namee.find('cBR8') >= 0 or namee.find('CMTS') >= 0 or namee.find('Anyang') >= 0 or namee.find('RFGW') >= 0:
					if namee.find('6509') >= 0 or namee.find('4507') >= 0: #안양 hostname 예외 경우 
						rns_number += 1
					
					else:
						cmts_number += 1
					#continue

				#R&S 개수 세기 
				else:
					rns_number += 1

				log = data_crawling(f,namee)

				ws.cell(x,1).value = namee
				ws.cell(x,1).font = Font(size=15, bold=True)

				for i in log:
					x+=1
					ws.cell(x,1).value = i

				x+=1
				f.close()

	#부가 기능
	print("R&S 개수 : %d"%rns_number)
	print("10K 개수 : %d"%cmts_number)

#엑셀 저장하는 함수
def wbsave(wb):
	try:
		wb.save("반기 log 수집.xlsx")
		return wb
	except Exception as ex: # 에러 종류
		print('error', ex)
		return -1

#엑셀 문서를 닫는 함수
def wbclose(wb):
	try:
		wb.close()
		return wb

	except Exception as ex: # 에러 종류
		print('error', ex)
		return -1

#!! MAIN!!
if __name__ == "__main__":
	user_input_txt = input("로그 파일 경로:")
	wb = excelOpen()
	check_crawring(wb,user_input_txt)
	LOG_crawling(wb,user_input_txt)

	wbsave(wb)
	wbclose(wb)
