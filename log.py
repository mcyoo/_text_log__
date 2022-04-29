import openpyxl
from openpyxl.styles import PatternFill, Color, Font
import codecs
import os

# 매개변수 file 안에 hostname을 반환하는 함수
def name(file):
	name = 0
	file.seek(0)
	while True:
		line=file.readline()
		hop = line.count('#')
		if not line:
			break
		if hop == 0 or hop > 5: continue
		else :
			temp = line
			for i in range(0,hop):
				name+=temp.find('#')
				name+=1
				temp=line[name:]
				i+=1

			line=line[:name-1]
			break

	return line

#엑셀 문서를 만드는 함수
def excelOpen():
	try:
		wb = openpyxl.Workbook()
		return wb

	except Exception as ex: # 에러 종류
		print('error', ex)
		return -1

# 매개변수 file 안에 show log 를 찾아서, show log 안에 Aug, Sep 
def crawling(file,name):
	log = []
	log.clear()
	file.seek(0)
	while True:
		line=file.readline()
		tem=line.find('show log')
		if tem >= 0:
			while True:
				line1=file.readline()
				tem1=line1.find('Sep') # <-- 변경!!
				tem2=line1.find('Oct') # <-- 변경!!
				tem3=line1.find(name)
				if tem1 >= 0 or tem2 >= 0:
					log.append(line1)
				if tem3 >= 0:
					break
				if not line1:
					break

		if not line:
			break

	return log

#pathnote 경로 안에 .txt .log 로 끝나는 파일을 열어서 리스트에 문자열을 하나씩 엑셀에 저장하는 함수
def LOG_crawling(wb,pathnote):
	
	x = 1
	rns_number = 0
	cmts_number = 0

	ws = wb.active
	for (path, dir, files) in os.walk(pathnote):
	
		for filename in files:
			ext = os.path.splitext(filename)[-1]
			if ext == '.txt' or ext == '.log':
				print("%s/%s" % (path, filename))

				f = codecs.open("%s/%s" % (path, filename), 'r', "utf-8-sig", errors='ignore')

				#메모장에서 이름 가져오기
				namee = name(f)

				#10K 는 로그 저장 X 
				if namee.find('10K') >= 0 or namee.find('10k') >= 0 or namee.find('cBR8') >= 0 or namee.find('CMTS') >= 0 or namee.find('Anyang') >= 0 or namee.find('RFGW') >= 0:
					if namee.find('6509') >= 0 or namee.find('4507') >= 0: #안양 hostname 예외 경우 
						rns_number += 1
					
					else:
						cmts_number += 1
						continue

				#R&S 개수 세기 
				else:
					rns_number += 1

				log = crawling(f,namee)

				ws.cell(x,1).value = namee
				ws.cell(x,1).font = Font(size=15, bold=True)

				for i in log:
					x+=1
					ws.cell(x,1).value = i

				x+=1

				f.close()

	#부가 기능
	print("R&S 개수 : %d"%rns_number)
	print("10K 개수 : %d"%cmts_number)

	
def wbsave(wb):
	try:
		wb.save("log 수집.xlsx")
		return wb
	except Exception as ex: # 에러 종류
		print('error', ex)
		return -1


def wbclose(wb):
	try:
		wb.close()
		return wb

	except Exception as ex: # 에러 종류
		print('error', ex)
		return -1


if __name__ == "__main__":
	user_input_path = input("로그 파일 경로:")
	wb = excelOpen()#excel 생성
	LOG_crawling(wb,user_input_path)#경로에서 로그 수집
	wbsave(wb)#excel 에 저장
	wbclose(wb)#메모리 반환

