import openpyxl
from openpyxl.styles import PatternFill, Color, Font
import codecs
import os

search_line_start = "↓↓↓ 찾으시는 단어 또는 문장을 한 줄씩 입력해주세요. (start) 예시 : 김치볶음밥 ↓↓↓"
search_line_end = "↑↑↑ 찾으시는 단어 또는 문장을 한 줄씩 입력해주세요. (end) ↑↑↑"

filter_line_start = "↓↓↓ 시작하는 단어 또는 문장 ~ 끝나는 단어 또는 문장을 입력해주세요. (start) 예시 : 김치볶음밥~간장밥 ↓↓↓"
filter_line_end = "↑↑↑ 시작하는 단어 또는 문장 ~ 끝나는 단어 또는 문장을 입력해주세요. (end) ↑↑↑"

user_file_name = 'save.txt'

# 매개변수 file 안에 hostname을 반환하는 함수
# ex) SCH_01#show run <- 에서 SCH_01 을 반환
# 이 프로그램에서 사용 안함


def name(file):
    name = 0
    file.seek(0)
    while True:
        line = file.readline()
        hop = line.count('#')
        if not line:
            break
        if hop == 0 or hop > 5:
            continue
        else:
            temp = line
            for i in range(0, hop):
                name += temp.find('#')
                name += 1
                temp = line[name:]
                i += 1

            line = line[:name-1]
            break

    return line


# 사용자가 입력한 파일 불러오기 (찾기)

def user_save_file1():
    with open(user_file_name, 'r') as file:
        lines = file.readlines()

    line_get = False
    user_search_lines = []

    for line in lines:
        if line.find(search_line_end) >= 0:
            line_get = False

        if line_get == True:
            user_search_lines.append(line)

        if line.find(search_line_start) >= 0:
            line_get = True

    return user_search_lines

# 사용자가 입력한 파일 불러오기 (필터)


def user_save_file2():
    with open(user_file_name, 'r') as file:
        lines = file.readlines()

    line_get = False
    user_filter_lines = []

    for line in lines:
        if line.find(filter_line_end) >= 0:
            line_get = False

        if line_get == True:
            user_filter_lines.append(line)

        if line.find(filter_line_start) >= 0:
            line_get = True

    return user_filter_lines


# 엑셀 문서를 만드는 함수
def excelOpen():
    try:
        wb = openpyxl.Workbook()

    except Exception as ex:  # 에러 종류
        print('엑셀 에러.. 코드 : ', ex)
        print('1. 엑셀이 설치가 되어있나요..?')
        print('2. 엑셀이 바로 열리나요?')
        print('3. 엑셀을 모두 닫고 해주세요')
        print('그래도 안되면 위에 에러 코드를 github에 코멘트 남겨주시면 업데이트 하겠습니다!')
        return -1
    return wb

# 파일에서 사용자가 입력한 단어 또는 문장을 찾는다.


def find_user_search(file, user_search_lines):
    log = []
    file.seek(0)
    lines = file.readlines()

    for line in lines:
        for user_line in user_search_lines:
            if line.find(user_line) >= 0:
                log.append(line)

    return log

# 파일에서 사용자가 입력한 단어 또는 문장에 시작 과 끝을 필터링 한다.


def find_user_filter(file, user_filter_lines):
    log = []
    log_temp = []
    file.seek(0)
    lines = file.readlines()

    for filter_line in user_filter_lines:
        line_get = False
        log_temp.clear()

        try:
            start_line, end_line = filter_line.split('~')
        except:
            continue
        for line in lines:
            if line.find(start_line) >= 0:
                line_get = True

            if line_get == True:
                log_temp.append(line)

            if line.find(end_line) >= 0:
                line_get = False
                log += log_temp
                break
    return log

# 파일 리스트에  있는 파일을 하나씩 열면서 사용자가 입력한 정보를 찾아엑셀 시트에 저장한다


def wb_create_sheets(wb, file_list):

    # 사용자가 저장한 텍스트 파일 가져오기
    user_search_lines = user_save_file1()
    user_filter_lines = user_save_file2()

    # /n 키 삭제
    user_search_lines = enter_key_del(user_search_lines)
    user_filter_lines = enter_key_del(user_filter_lines)

    excel_index_log1 = 1
    excel_index_log2 = 1

    ws_search = wb.create_sheet('수집 데이터 search')
    ws_filter = wb.create_sheet('수집 데이터 filter')

    for path, filename in file_list:
        f = codecs.open("%s/%s" % (path, filename), 'r',
                        "utf-8-sig", errors='ignore')

        log1 = find_user_search(f, user_search_lines)
        log2 = find_user_filter(f, user_filter_lines)

        # /n 키 삭제
        log1 = enter_key_del(log1)
        log2 = enter_key_del(log2)

        # log1 | find_user_search | user_search_lines | 수집 데이터 search
        ws_search.cell(excel_index_log1, 1).value = filename
        ws_search.cell(excel_index_log1, 1).font = Font(
            size=20, bold=True)

        for line in log1:
            excel_index_log1 += 1
            ws_search.cell(excel_index_log1, 1).value = line
        excel_index_log1 += 1

        # log2 | find_user_filter | user_filter_lines | 수집 데이터 filter
        ws_filter.cell(excel_index_log2, 1).value = filename
        ws_filter.cell(excel_index_log2, 1).font = Font(
            size=20, bold=True)

        for line in log2:
            excel_index_log2 += 1
            ws_filter.cell(excel_index_log2, 1).value = line
        excel_index_log2 += 1

        f.close()


# pathnote 경로 안에 .txt .log 로 끝나는 파일을 열어서 리스트에 문자열을 하나씩 엑셀에 저장하는 함수
def crawling(pathnote):
    file_list = []
    for (path, dir, files) in os.walk(pathnote):
        for filename in files:
            ext = os.path.splitext(filename)[-1]
            if ext == '.txt' or ext == '.log':
                print("%s/%s" % (path, filename))
                file_list.append((path, filename))
    return file_list


# 엑셀 저장하는 함수
def wbsave(wb):
    try:
        wb.save("수집된 엑셀파일.xlsx")
        print('엑셀 저장 완료(입력하신 파일 경로에 파일이 저장되었습니다. 엑셀 이름 : 수집된 엑셀파일.xlsx')
    except Exception as ex:  # 에러 종류
        print('엑셀 저장 에러(엑셀 프로그램을 모두 끄고 다시 실행해주세요.)', ex)
        return -1
    return wb

# 입력받는 input_list 안에 \n 키 삭제


def enter_key_del(input_list):
    output_list = []
    for x in input_list:
        if x.find('\n') >= 0:
            x = x[:-1]
        output_list.append(x)
    return output_list

# 엑셀 문서를 닫는 함수


def wbclose(wb):
    try:
        wb.close()
    except Exception as ex:  # 에러 종류
        print('엑셀 프로세스 에러(엑셀 프로그램을 모두 끄고 다시 실행해주세요.)', ex)
        return -1
    return wb


#!! MAIN!!
if __name__ == "__main__":
    user_input_txt = input("로그 파일 경로:")
    result = crawling(user_input_txt)

    wb = excelOpen()
    wb_create_sheets(wb, result)
    wbsave(wb)
    wbclose(wb)
